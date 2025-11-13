import pandas as pd
from pathlib import Path
import time
import win32com.client as win32
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Display process to user
print("🔄 Starting Sales Data Consolidation...")

# List of source files
source_files = [f"Coffee Shop {i}.xlsx" for i in range(1, 6)]

# Verify files exist
for file in source_files:
    if not Path(file).exists():
        print(f"❌ File not found: {file}")
        exit(1)

# Step 1: Read and merge all Excel files
print("📂 Reading Excel files...")
excel = win32.gencache.EnsureDispatch('Excel.Application')
excel.Visible = True
dataframes = []
for file in source_files:
    print(f"   ➜ Reading {file}...")
    df = pd.read_excel(file)
    dataframes.append(df)
    wb = excel.Workbooks.Open(str(Path(file).resolve()))
    time.sleep(5)
    wb.Close(SaveChanges=False)
    

# Merge all data into one DataFrame
print("🧩 Merging all data...")
merged_df = pd.concat(dataframes, ignore_index=True)

# Step 2: Create Pivot Table
print("📊 Creating pivot table (Total Income by Item)...")
pivot_df = merged_df.pivot_table(
    index='Item',
    values='Total Income ($)',
    aggfunc='sum'
).reset_index()

# Step 3: Write merged data and pivot table to Excel
output_file = "SalesResult.xlsx"
print(f"💾 Writing consolidated data to {output_file}...")

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    merged_df.to_excel(writer, index=False, sheet_name='MergedData')
    pivot_df.to_excel(writer, index=False, sheet_name='PivotSummary')

# Step 4: Display completion summary
print("✅ Sales data consolidation complete!")
print(f"   ➜ Merged sheet: {len(merged_df)} records")
print(f"   ➜ Pivot summary: {len(pivot_df)} unique items")
print(f"📘 Output saved as '{output_file}'")

# Optional: Display preview of data in console
print("\n--- Preview of merged data ---")
print(merged_df.head())

print("\n--- Preview of pivot summary ---")
print(pivot_df.head())

excel = win32.gencache.EnsureDispatch('Excel.Application')
excel.Visible = True
wb = excel.Workbooks.Open(str(Path(output_file).resolve()))
ws = wb.Sheets("MergedData")

# Format header row (Row 1)
header_range = ws.Range("A1:D1")  # adjust columns as needed
header_range.Font.Bold = True
time.sleep(2)
header_range.Interior.ColorIndex = 10  # green fill
time.sleep(2)
header_range.Font.ColorIndex = 2      # white text
header_range.HorizontalAlignment = -4108  # center
time.sleep(2)

# Auto fit columns
ws.Columns.AutoFit()

# Pivot sheet header formatting
pivot_ws = wb.Sheets("PivotSummary")
pivot_ws.Range("A1:B1").Font.Bold = True
pivot_ws.Range("A1:B1").Interior.ColorIndex = 10
pivot_ws.Range("A1:B1").Font.ColorIndex = 2
pivot_ws.Columns.AutoFit()

print("🎨 Formatting complete (via Excel).")

time.sleep(10)
wb.Close(SaveChanges=True)
excel.Quit()