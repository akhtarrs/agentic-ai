import pandas as pd
from pathlib import Path
import time
import win32com.client as win32
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Display process to user
print("🔄 Starting Sales Data Consolidation...")

# List of source files
source_files = [f"Coffee Shop {i}.xlsx" for i in range(1, 6)]

# Verify files exist
for file in source_files:
    if not Path(file).exists():
        print(f"❌ File not found: {file}")
        exit(1)

# Step 1: Read and merge all Excel files
print("📂 Reading Excel files...")
excel = win32.gencache.EnsureDispatch('Excel.Application')
excel.Visible = True
dataframes = []
for file in source_files:
    print(f"   ➜ Reading {file}...")
    df = pd.read_excel(file)
    dataframes.append(df)
    wb = excel.Workbooks.Open(str(Path(file).resolve()))
    time.sleep(5)
    wb.Close(SaveChanges=False)
    

# Merge all data into one DataFrame
print("🧩 Merging all data...")
merged_df = pd.concat(dataframes, ignore_index=True)

# Step 2: Create Pivot Table
print("📊 Creating pivot table (Total Income by Item)...")
pivot_df = merged_df.pivot_table(
    index='Item',
    values='Total Income ($)',
    aggfunc='sum'
).reset_index()

# Step 3: Write merged data and pivot table to Excel
output_file = "SalesResult.xlsx"
print(f"💾 Writing consolidated data to {output_file}...")

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    merged_df.to_excel(writer, index=False, sheet_name='MergedData')
    pivot_df.to_excel(writer, index=False, sheet_name='PivotSummary')

# Step 4: Display completion summary
print("✅ Sales data consolidation complete!")
print(f"   ➜ Merged sheet: {len(merged_df)} records")
print(f"   ➜ Pivot summary: {len(pivot_df)} unique items")
print(f"📘 Output saved as '{output_file}'")

# Optional: Display preview of data in console
print("\n--- Preview of merged data ---")
print(merged_df.head())

print("\n--- Preview of pivot summary ---")
print(pivot_df.head())

sr = excel.Workbooks.Open(str(Path(output_file).resolve()))
print("✅ Data written successfully. Now applying formatting...")

# Step 4: Format Excel file
wb = load_workbook(output_file)
ws = wb["MergedData"]

# Define header style
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")
header_alignment = Alignment(horizontal="center", vertical="center")

# Apply header formatting
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Auto-fit column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column
    for cell in col:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = max_length + 2
    ws.column_dimensions[get_column_letter(column)].width = adjusted_width

# Optional: format PivotSummary headers too
ws_pivot = wb["PivotSummary"]
for cell in ws_pivot[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Auto-fit column widths
for col in ws_pivot.columns:
    max_length = 0
    column = col[0].column
    for cell in col:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = max_length + 2
    ws_pivot.column_dimensions[get_column_letter(column)].width = adjusted_width


# Save formatted file
wb.save(output_file)
wb.close()

print("🎨 Formatting complete.")

time.sleep(10)
sr.Close(SaveChanges=False)
excel.Quit()
